from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func
from fastapi.responses import FileResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER
import datetime
import os

from config.database import get_db
from Model.Report import Report

router = APIRouter()


# ✅ GET tous les reports
@router.get("/reports")
def get_all_reports(db: Session = Depends(get_db)):
    reports = db.query(Report).order_by(Report.id.desc()).all()
    return [
        {
            "id": r.id,
            "total_locations": r.total_locations,
            "revenue_total": r.revenue_total,
            "prix_moyen": r.prix_moyen,
            "date_creation": str(r.date_creation)
        }
        for r in reports
    ]


# ✅ GET stats globales
@router.get("/reports/stats")
def get_stats(db: Session = Depends(get_db)):
    total_reports = db.query(func.count(Report.id)).scalar()
    total_revenue = db.query(func.sum(Report.revenue_total)).scalar() or 0
    average_price = db.query(func.avg(Report.prix_moyen)).scalar() or 0
    return {
        "total_reports": total_reports,
        "total_revenue": round(float(total_revenue), 2),
        "average_price": round(float(average_price), 2)
    }


# ✅ GET dashboard
@router.get("/reports/dashboard")
def get_dashboard(db: Session = Depends(get_db)):
    report = db.query(Report).order_by(Report.id.desc()).first()
    if not report:
        return {"total_locations": 0, "total_revenue": 0, "average_price": 0}
    return {
        "total_locations": report.total_locations,
        "total_revenue": report.revenue_total,
        "average_price": round(float(report.prix_moyen), 2)
    }


# ✅ POST générer un nouveau rapport
@router.post("/reports/generate")
def generate_report(db: Session = Depends(get_db)):
    total_locations = db.query(func.sum(Report.total_locations)).scalar() or 0
    revenue_total   = db.query(func.sum(Report.revenue_total)).scalar() or 0
    total_reports   = db.query(func.count(Report.id)).scalar() or 1
    prix_moyen      = revenue_total / total_reports

    new_report = Report(
        total_locations=int(total_locations),
        revenue_total=round(float(revenue_total), 2),
        prix_moyen=round(float(prix_moyen), 2)
    )
    db.add(new_report)
    db.commit()
    db.refresh(new_report)
    return {
        "message": "Rapport cree avec succes",
        "data": {
            "id": new_report.id,
            "total_locations": new_report.total_locations,
            "revenue_total": new_report.revenue_total,
            "prix_moyen": new_report.prix_moyen,
            "date_creation": str(new_report.date_creation)
        }
    }


# ✅ GET export PDF
@router.get("/reports/export/pdf")
def export_pdf(db: Session = Depends(get_db)):
    report = db.query(Report).order_by(Report.id.desc()).first()
    if not report:
        return {"message": "No report found"}

    file_path = "/tmp/report.pdf"
    doc = SimpleDocTemplate(file_path, pagesize=(210*mm, 297*mm), rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)

    BLUE_DARK  = colors.HexColor("#1a237e")
    BLUE_MID   = colors.HexColor("#3f51b5")
    BLUE_LIGHT = colors.HexColor("#e8eaf6")
    GREEN      = colors.HexColor("#4CAF50")
    ORANGE     = colors.HexColor("#FF9800")
    GREY       = colors.HexColor("#757575")
    WHITE      = colors.white

    title_style       = ParagraphStyle("title_style",    fontSize=22, textColor=WHITE,     alignment=TA_CENTER, fontName="Helvetica-Bold")
    subtitle_style    = ParagraphStyle("subtitle_style", fontSize=10, textColor=colors.HexColor("#c5cae9"), alignment=TA_CENTER, fontName="Helvetica")
    section_style     = ParagraphStyle("section_style",  fontSize=13, textColor=BLUE_DARK, fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    footer_style      = ParagraphStyle("footer_style",   fontSize=8,  textColor=GREY,      alignment=TA_CENTER, fontName="Helvetica-Oblique")
    card_label        = ParagraphStyle("card_label",     fontSize=9,  textColor=GREY,      fontName="Helvetica", alignment=TA_CENTER)
    card_value        = ParagraphStyle("card_value",     fontSize=18, textColor=BLUE_DARK, fontName="Helvetica-Bold", alignment=TA_CENTER)
    card_value_green  = ParagraphStyle("cvg",            fontSize=18, textColor=GREEN,     fontName="Helvetica-Bold", alignment=TA_CENTER)
    card_value_orange = ParagraphStyle("cvo",            fontSize=18, textColor=ORANGE,    fontName="Helvetica-Bold", alignment=TA_CENTER)
    label_style       = ParagraphStyle("lbl",            fontSize=10, textColor=BLUE_DARK, fontName="Helvetica-Bold")
    value_style       = ParagraphStyle("val",            fontSize=10, textColor=colors.HexColor("#424242"), fontName="Helvetica")

    elements = []

    # HEADER AVEC LOGO
   # HEADER AVEC LOGO
    logo_path = "/app/logo.png"

    if os.path.exists(logo_path):
        logo = Image(logo_path, width=20*mm, height=20*mm)
    else:
        logo = Paragraph("", title_style)

    title_para    = Paragraph("ERP Report Service", title_style)
    subtitle_para = Paragraph("Location - Vehicules - Equipe", subtitle_style)

    # Logo à gauche | Titre + Sous-titre à droite
    header_table = Table(
        [[logo, title_para],
         [""  , subtitle_para]],
        colWidths=[28*mm, 140*mm]
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), BLUE_DARK),
        ("VALIGN",        (0, 0), (0, -1),  "MIDDLE"),
        ("VALIGN",        (1, 0), (1, 0),   "BOTTOM"),
        ("VALIGN",        (1, 1), (1, 1),   "TOP"),
        ("SPAN",          (0, 0), (0, 1)),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
    ]))
    elements.append(header_table)

    sub_table = Table([[Paragraph("Generated Report - ERP Car Rental System", subtitle_style)]], colWidths=[170*mm])
    sub_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), BLUE_MID),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(sub_table)
    elements.append(Spacer(1, 16))

    # CARDS
    elements.append(Paragraph("Report Summary", section_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=BLUE_MID, spaceAfter=10))

    card_unit        = ParagraphStyle("card_unit", fontSize=9, textColor=GREY,   fontName="Helvetica", alignment=TA_CENTER)
    card_unit_green  = ParagraphStyle("cug",       fontSize=9, textColor=GREEN,  fontName="Helvetica", alignment=TA_CENTER)
    card_unit_orange = ParagraphStyle("cuo",       fontSize=9, textColor=ORANGE, fontName="Helvetica", alignment=TA_CENTER)

    cards_data = [[
        Table([
            [Paragraph("Total Locations", card_label)],
            [Paragraph(str(report.total_locations), card_value)],
            [Paragraph("locations", card_unit)]
        ], colWidths=[50*mm]),
        Table([
            [Paragraph("Total Revenue", card_label)],
            [Paragraph(f"{report.revenue_total:,.2f}", card_value_green)],
            [Paragraph("EUR", card_unit_green)]
        ], colWidths=[50*mm]),
        Table([
            [Paragraph("Average Price", card_label)],
            [Paragraph(f"{report.prix_moyen:,.2f}", card_value_orange)],
            [Paragraph("EUR", card_unit_orange)]
        ], colWidths=[50*mm]),
    ]]

    cards_table = Table(cards_data, colWidths=[56*mm, 56*mm, 56*mm])
    cards_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (0, 0), BLUE_LIGHT),
        ("BACKGROUND",    (1, 0), (1, 0), colors.HexColor("#e8f5e9")),
        ("BACKGROUND",    (2, 0), (2, 0), colors.HexColor("#fff3e0")),
        ("TOPPADDING",    (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
    ]))
    elements.append(cards_table)
    elements.append(Spacer(1, 20))

    # DETAILS
    elements.append(Paragraph("Report Details", section_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=BLUE_MID, spaceAfter=10))

    details = [
        [Paragraph("Field",           label_style), Paragraph("Value", label_style)],
        [Paragraph("Report ID",       value_style), Paragraph(str(report.id), value_style)],
        [Paragraph("Total Locations", value_style), Paragraph(str(report.total_locations), value_style)],
        [Paragraph("Total Revenue",   value_style), Paragraph(f"{report.revenue_total:,.2f} EUR", value_style)],
        [Paragraph("Average Price",   value_style), Paragraph(f"{report.prix_moyen:,.2f} EUR", value_style)],
        [Paragraph("Date Generated",  value_style), Paragraph(str(report.date_creation), value_style)],
    ]
    detail_table = Table(details, colWidths=[60*mm, 108*mm])
    detail_table.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  BLUE_MID),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  WHITE),
        ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BLUE_LIGHT, WHITE]),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.HexColor("#bdbdbd")),
        ("TOPPADDING",     (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 9),
        ("LEFTPADDING",    (0, 0), (-1, -1), 12),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 12),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 30))

    # FOOTER
    elements.append(HRFlowable(width="100%", thickness=1, color=BLUE_MID, spaceAfter=8))
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"Generated automatically by ERP Report Service  -  {now}", footer_style))

    doc.build(elements)
    return FileResponse(file_path, filename="report.pdf", media_type="application/pdf")


# ✅ GET export EXCEL
@router.get("/reports/export/excel")
def export_excel(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    reports = db.query(Report).order_by(Report.id.desc()).all()
    if not reports:
        return {"message": "No reports found"}

    wb = Workbook()
    ws = wb.active
    ws.title = "ERP Reports"

    BLUE_DARK = "1A237E"
    BLUE_MID  = "3F51B5"
    BLUE_LIGHT= "E8EAF6"
    GREEN     = "4CAF50"
    ORANGE    = "FF9800"
    WHITE     = "FFFFFF"
    GREY      = "F5F5F5"

    # TITRE
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "ERP Report Service - Location Vehicules Equipe"
    title_cell.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:E2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated Report - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub_cell.font = Font(name="Arial", size=10, italic=True, color=WHITE)
    sub_cell.fill = PatternFill("solid", fgColor=BLUE_MID)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.append([])

    # STATS GLOBALES
    ws.merge_cells("A4:E4")
    stats_title = ws["A4"]
    stats_title.value = "STATISTIQUES GLOBALES"
    stats_title.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    stats_title.fill = PatternFill("solid", fgColor=BLUE_MID)
    stats_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22

    total_revenue = sum(r.revenue_total for r in reports)
    avg_price     = total_revenue / len(reports) if reports else 0
    total_locs    = sum(r.total_locations for r in reports)

    stats_data = [
        ["Total Rapports", len(reports), "", "Total Locations", total_locs],
        ["Revenue Total", f"{total_revenue:,.2f} EUR", "", "Prix Moyen", f"{avg_price:,.2f} EUR"],
    ]

    thin = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_data in stats_data:
        ws.append(row_data)
        row = ws.max_row
        ws.row_dimensions[row].height = 20
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
            cell.font = Font(name="Arial", size=10, bold=(col in [1, 4]))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    ws.append([])

    # HEADER TABLEAU
    headers = ["ID", "Total Locations", "Revenue Total (EUR)", "Prix Moyen (EUR)", "Date Creation"]
    ws.append(headers)
    header_row = ws.max_row
    ws.row_dimensions[header_row].height = 22

    white_side = Side(style="thin", color=WHITE)
    white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = white_border

    # DONNÉES
    for i, r in enumerate(reports):
        ws.append([r.id, r.total_locations, round(r.revenue_total, 2), round(r.prix_moyen, 2), str(r.date_creation)])
        row = ws.max_row
        ws.row_dimensions[row].height = 18
        bg = GREY if i % 2 == 0 else WHITE

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if col == 3:
                cell.font = Font(name="Arial", size=10, color=GREEN, bold=True)
            if col == 4:
                cell.font = Font(name="Arial", size=10, color=ORANGE, bold=True)

    # LARGEUR COLONNES
    for i, width in enumerate([8, 18, 22, 20, 25], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # FOOTER
    ws.append([])
    ws.append(["Generated automatically by ERP Report Service"])
    footer_row = ws.max_row
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    footer_cell = ws[f"A{footer_row}"]
    footer_cell.font = Font(name="Arial", size=8, italic=True, color="757575")
    footer_cell.alignment = Alignment(horizontal="center")

    file_path = "/tmp/report.xlsx"
    wb.save(file_path)

    return FileResponse(
        file_path,
        filename="report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )